"""Bounded native DOCX/XLSX manual extraction for candidate rules."""

from __future__ import annotations

import hashlib
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook

from brp.adapters.contracts import (
    AdapterDiagnostic,
    CandidateDecision,
    ExtractionBatch,
    Source,
    SourceSnapshot,
    UnmappableItem,
)
from brp.ir.models import DecisionContent

HEADERS = {
    "decision_key",
    "decision_name",
    "rule_id",
    "input_name",
    "input_type",
    "operator",
    "operand_value",
    "output_name",
    "output_type",
    "output_value",
}


class ManualDocumentAdapter:
    name = "docs-manual"
    capability_version = "source-adapter/v1"

    def __init__(self, paths: list[Path], *, confidence: float = 0.35) -> None:
        if not 0 <= confidence < 0.5:
            raise ValueError("manual adapter default confidence must remain below 0.5")
        self.paths = [path.resolve() for path in paths]
        self.confidence = confidence

    def discover(self, site_config: object) -> list[Source]:
        del site_config
        sources: list[Source] = []
        for path in sorted(self.paths):
            suffix = path.suffix.lower()
            if suffix not in {".docx", ".xlsx"}:
                continue
            sources.append(
                Source(
                    source_id=f"manual:{path.name}",
                    kind="manual-document",
                    locator={"path": str(path), "format": suffix[1:]},
                )
            )
        return sources

    def extract(self, source: Source) -> ExtractionBatch:
        path = Path(str(source.locator["path"]))
        content_hash = hashlib.sha256(path.read_bytes()).hexdigest()
        if path.suffix.lower() == ".xlsx":
            decisions, unmappable = self._extract_xlsx(path, content_hash)
        elif path.suffix.lower() == ".docx":
            decisions, unmappable = [], self._extract_docx(path, content_hash)
        else:
            raise ValueError(f"unsupported manual format: {path.suffix}")
        return ExtractionBatch(
            adapter=self.name,
            decisions=decisions,
            unmappable=unmappable,
            diagnostics=[
                AdapterDiagnostic(
                    level="INFO",
                    code="NATIVE_MANUAL_PARSED",
                    message=f"parsed {path.name} with its native document library",
                )
            ],
            source_snapshot=SourceSnapshot(
                source_id=source.source_id,
                revision=content_hash,
                content_hash=content_hash,
                captured_at=datetime.now(UTC),
            ),
        )

    def _extract_xlsx(
        self, path: Path, content_hash: str
    ) -> tuple[list[CandidateDecision], list[UnmappableItem]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        decisions: list[CandidateDecision] = []
        unmappable: list[UnmappableItem] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            if not HEADERS.issubset(headers):
                for number, row in enumerate(rows[1:], 2):
                    raw = " | ".join(str(value) for value in row if value not in (None, ""))
                    if raw:
                        unmappable.append(
                            _unmappable(path, content_hash, raw, sheet.title, f"A{number}")
                        )
                continue
            records = [dict(zip(headers, row, strict=False)) for row in rows[1:]]
            grouped: dict[str, list[tuple[int, dict[str, Any]]]] = defaultdict(list)
            for number, record in enumerate(records, 2):
                key = str(record.get("decision_key") or "").strip()
                if not key:
                    raw = " | ".join(
                        str(value) for value in record.values() if value not in (None, "")
                    )
                    unmappable.append(
                        _unmappable(
                            path,
                            content_hash,
                            raw or "blank decision key",
                            sheet.title,
                            f"A{number}",
                        )
                    )
                else:
                    grouped[key].append((number, record))
            decisions.extend(
                self._decision(path, content_hash, sheet.title, key, values)
                for key, values in sorted(grouped.items())
            )
        workbook.close()
        return decisions, unmappable

    def _decision(
        self,
        path: Path,
        content_hash: str,
        sheet: str,
        key: str,
        values: list[tuple[int, dict[str, Any]]],
    ) -> CandidateDecision:
        first = values[0][1]
        input_name = str(first["input_name"])
        output_name = str(first["output_name"])
        document = {
            "decisionId": key,
            "decisionName": str(first["decision_name"]),
            "profile": "RULE_IR_V1",
            "schemaVersion": 1,
            "programContexts": [
                {"programId": "MANUAL", "kind": "SERVICE", "entryPoint": path.name}
            ],
            "hitPolicy": "FIRST",
            "inputs": [{"name": input_name, "type": str(first["input_type"]), "required": True}],
            "outputs": [{"name": output_name, "type": str(first["output_type"])}],
            "defaultOutput": {output_name: _default(str(first["output_type"]))},
            "rules": [
                {
                    "ruleId": str(record["rule_id"]),
                    "when": {
                        "all": [
                            {
                                "left": {"kind": "INPUT", "name": input_name},
                                "operator": str(record["operator"]),
                                "right": {
                                    "kind": "LITERAL",
                                    "value": _typed(
                                        record["operand_value"], str(first["input_type"])
                                    ),
                                },
                            }
                        ]
                    },
                    "then": [
                        {
                            "output": output_name,
                            "value": _typed(record["output_value"], str(first["output_type"])),
                        }
                    ],
                    "origin": "EXTRACTED",
                    "sourceReferences": [
                        {
                            "type": "MANUAL_DOC",
                            "documentId": path.name,
                            "revision": content_hash,
                            "sheet": sheet,
                            "section": str(record.get("section") or key),
                            "cellRange": f"A{number}:K{number}",
                        }
                    ],
                    "confidence": self.confidence,
                }
                for number, record in values
            ],
        }
        return CandidateDecision(decision_key=key, content=DecisionContent.model_validate(document))

    def _extract_docx(self, path: Path, content_hash: str) -> list[UnmappableItem]:
        document = Document(str(path))
        section = "Document"
        items: list[UnmappableItem] = []
        for paragraph in document.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            if paragraph.style is not None and paragraph.style.name.startswith("Heading"):
                section = text
                continue
            items.append(_unmappable(path, content_hash, text, None, None, section))
        return items


def _unmappable(
    path: Path,
    revision: str,
    raw: str,
    sheet: str | None,
    cell: str | None,
    section: str | None = None,
) -> UnmappableItem:
    provenance = {"documentId": path.name, "revision": revision}
    if sheet:
        provenance["sheet"] = sheet
    if cell:
        provenance["cellRange"] = cell
    if section:
        provenance["section"] = section
    return UnmappableItem(
        reason_code="AMBIGUOUS_MANUAL_TEXT",
        raw_fragment=raw,
        provenance=provenance,
    )


def _typed(value: object, type_name: str) -> object:
    if type_name == "integer":
        return int(str(value))
    if type_name == "decimal":
        return float(str(value))
    if type_name == "boolean":
        return bool(value)
    return str(value)


def _default(type_name: str) -> object:
    return {"integer": 0, "decimal": 0.0, "boolean": False}.get(type_name, "")
